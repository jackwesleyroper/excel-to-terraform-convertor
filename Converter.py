import os
import sys
import openpyxl
from openpyxl import load_workbook
i = 0

start = 'FF686996'
normal = 'FFBDBDD1'
end = 'FF00CC99'

file = sys.argv[1]
print("Loading:" + file)
workbook = load_workbook(filename=file)
sheet = workbook.active
rows = list(sheet.iter_rows(min_col=1, max_col=3))

def autovars():
        global i
        global rows
        f.write(cells[1].value + " = {\n")
        #start count within sheet + 1 to skip header
        j = i+1
        #start count within block
        l = 0
        for block in list(rows)[i+1:]:
                if block[0].fill.start_color.index == start:
                        #if the block 0 colour is dark purple
                        if l > 1:
                                #if the count within a section is greater than 1
                                if block[1].value == "^":
                                        #if ^ is in block 1 then create double tabbed }
                                        f.write("\t\t}\n")
                                else:
                                        #otherwise assume single tabbed }
                                        f.write("\t}\n")
                        if not block[2].value:
                                #if part of map within a map
                                print(fileout + ": " + block[0].value)
                                f.write("\t" + block[0].value + " = {\n")
                        else:
                                if block[0].value:
                                        #if there is a name in block 0 then print it and write starting block
                                        print(fileout + ": " + block[0].value)
                                        f.write("\t" + block[0].value + " = {\n")
                                #otherwise write double tabbed starting header
                                f.write("\t\t" + block[2].value + " = {\n")

                if block[0].fill.start_color.index == normal:
                        #if the block 0 colour is light purple
                        #convert input to str
                        value = str(block[1].value)
                        if block[2].value == "^":
                                #if the block 1 value is ^ triple tab the output
                                if not block[0].value:
                                        #if part of a list
                                        f.write("\"" + value + "\", ")
                                elif block[0].value == "]":
                                        #if list stop
                                        f.seek(0,2)
                                        size=f.tell()
                                        f.truncate(size-2)
                                        f.write(block[0].value + "\n")
                                else:
                                        if value in ['True', 'False']:
                                                #if bool value then leave out "" marks
                                                f.write("\t\t\t" + block[0].value + " = " + value + "\n")
                                        elif value == "[":
                                                #if list start
                                                f.write("\t\t\t" + block[0].value + " = " + value)
                                        else:
                                                #otherwise print as normal
                                                f.write("\t\t\t" + block[0].value + " = \"" + value + "\"\n")
                        else:
                                #otherwise 
                                if not block[0].value:
                                        #if part of a list
                                        f.write("\"" + value + "\", ")
                                elif block[0].value == "]":
                                        #if list stop
                                        f.seek(0,2)
                                        size=f.tell()
                                        f.truncate(size-2)
                                        f.write(block[0].value + "\n")
                                else:
                                        if value in ['True', 'False']:
                                                #if bool value then leave out "" marks
                                                f.write("\t\t" + block[0].value + " = " + value + "\n")
                                        elif value == "[":
                                                #if list start
                                                f.write("\t\t" + block[0].value + " = " + value)
                                        else:
                                                #otherwise print as normal
                                                f.write("\t\t" + block[0].value + " = \"" + value + "\"\n")

                if block[0].fill.start_color.index == end:
                        #if the block 0 colour is mint
                        if block[2].value == "^":
                                #if block 2 is ^ then move down from triple tab
                                f.write("\t\t}\n")
                                f.write("\t}\n")
                                f.write("}\n")
                        else:
                                #otherwise assume double tab
                                f.write("\t}\n")
                                f.write("}\n")
                        i=j
                        break
                l+=1
                j+=1
def regular():
        global i
        global rows
        #write standard variable starter and report to console
        print(fileout + ": " + cells[1].value)
        f.write("variable \"" + cells[1].value + "\" {\n")
        j = i+1
        for block in list(rows)[i+1:]:
                if block[0].fill.start_color.index == normal:
                        #if the block 0 colour is light purple
                        value = str(block[1].value)
                        if block[2].value == "^":
                                #if the block 1 value is ^ double tab the output and create a map
                                if not block[0].value:
                                        #if part of a list
                                        f.write("\"" + value + "\", ")
                                elif block[0].value == "]":
                                        #if list stop
                                        f.seek(0,2)
                                        size=f.tell()
                                        f.truncate(size-2)
                                        f.write(block[0].value + "\n")
                                else:
                                        if value in ['True', 'False']:
                                                #if bool value then leave out "" marks
                                                f.write("\t\t" + block[0].value + " = " + value + "\n")
                                        elif value == "[":
                                                #if list start
                                                f.write("\t\t" + block[0].value + " = " + value)
                                        else:
                                                #otherwise print as normal
                                                f.write("\t\t" + block[0].value + " = \"" + value + "\"\n")
                        else:
                                #otherwise single tab the output
                                if not block[0].value:
                                        #if part of a list
                                        f.write("\"" + value + "\", ")
                                elif block[0].value == "]":
                                        #if list stop
                                        f.seek(0,2)
                                        size=f.tell()
                                        f.truncate(size-2)
                                        f.write(block[0].value + "\n")
                                else:
                                        if value in ['True', 'False']:
                                                #if bool value then leave out "" marks
                                                f.write("\t" + block[0].value + " = " + value + "\n")
                                        elif value == "[":
                                                #if list start
                                                f.write("\t" + block[0].value + " = " + value)
                                        else:
                                                #otherwise print as normal
                                                f.write("\t" + block[0].value + " = \"" + value + "\"\n")

                if block[0].fill.start_color.index == end:
                        #if the block 0 colour is mint
                        f.write("}\n")
                        i=j
                        break
                j+=1
while i < len(rows):

        cells = rows[i]

        if cells[0].fill.start_color.index == end:
                #if the block 0 colour is mint
                f.close()
                break

        if cells[0].fill.start_color.index == start:
                #if the block 0 colour is dark purple
                fileout = cells[0].value            
                #define file name from cells 0 value    
                f = open(fileout, "a")
                #create file from value

                if cells[2].value == "autovars":
                        #select map variable
                        autovars()

                if cells[2].value == "regular":
                        #select regular variable
                        regular()
        i+=1